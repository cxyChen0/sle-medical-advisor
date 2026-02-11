"""
检查单解析服务
基于现有的parse_reports.py脚本，适配FastAPI的UploadFile格式
"""

from typing import Dict, List, Optional
import re
import json
from datetime import datetime
from pathlib import Path
from fastapi import UploadFile
import tempfile
import os

from app.services.normalization_service import normalize_medical_terms
from app.config.settings import settings
from app.services.image_optimization_service import image_optimization_service

# 全局OCR对象缓存
_ocr_instance = None
_ocr_lock = None

def get_ocr_instance():
    """获取全局OCR实例，避免重复初始化"""
    global _ocr_instance, _ocr_lock
    
    if _ocr_instance is None:
        import threading
        _ocr_lock = _ocr_lock or threading.Lock()
        
        with _ocr_lock:
            if _ocr_instance is None:
                from cnocr import CnOcr
                model_dir = os.path.join(os.getcwd(), 'cnocr_models')
                os.makedirs(model_dir, exist_ok=True)
                os.environ['CNOCR_HOME'] = model_dir
                
                # 使用更快的模型配置
                _ocr_instance = CnOcr(
                    model_name='densenet_lite_136',  # 使用更快的模型
                    model_dir=model_dir,
                    rec_model_fp16=True,  # 使用FP16加速
                )
                print("OCR instance initialized (densenet_lite_136 model)")
    
    return _ocr_instance

from app.services.ai_semantic_service import ai_semantic_service, parse_report_with_ai


async def parse_report(file: UploadFile, report_type: str) -> Dict:
    """
    解析检查单
    
    Args:
        file: 上传的检查单文件
        report_type: 报告类型 (lab/pathology)
        
    Returns:
        解析结果字典
    """
    # 保存临时文件
    with tempfile.NamedTemporaryFile(delete=False, suffix=Path(file.filename).suffix) as temp_file:
        content = await file.read()
        temp_file.write(content)
        temp_file_path = temp_file.name
    
    try:
        # 提取文本
        text = extract_text_from_file(temp_file_path)
        print(f"Extracted text length: {len(text)}")
        
        # 解析指标
        indicators = await parse_indicators(text, report_type)
        print(f"Parsed indicators: {len(indicators)}")
        
        # 尝试提取检查日期
        date_match = re.search(r'(\d{4}[-/年]\d{1,2}[-/月]\d{1,2}日?)', text)
        report_date = date_match.group(1) if date_match else datetime.now().strftime('%Y-%m-%d')
        print(f"Report date: {report_date}")
        
        # 尝试提取患者ID
        patient_id_match = re.search(r'(患者ID|ID|就诊号|病历号).*?[:：]\s*([\w\d-]+)', text)
        patient_id = patient_id_match.group(2) if patient_id_match else None
        print(f"Patient ID: {patient_id}")
        
        normalized_terms = []
        normalized_indicators = []
        
        # 直接使用AI返回的归一化结果，不再进行额外的归一化调用
        if indicators:
            # AI已经在解析时完成了归一化，直接使用结果
            for indicator in indicators:
                # 确保归一化字段存在
                if 'normalized_name' not in indicator:
                    indicator['normalized_name'] = indicator.get('name', '')
                if 'normalization_confidence' not in indicator:
                    indicator['normalization_confidence'] = 1.0
                
                # 添加到归一化结果列表
                normalized_terms.append({
                    'original': indicator.get('name', ''),
                    'normalized': indicator.get('normalized_name', ''),
                    'confidence': indicator.get('normalization_confidence', 1.0)
                })
                
                normalized_indicators.append(indicator)
        else:
            normalized_indicators = indicators
        
        result = {
            'patient_id': patient_id,
            'report_date': report_date,
            'report_type': report_type,
            'indicators': normalized_indicators,
            'normalization_results': normalized_terms,
            'original_text': text  # 添加原始文本
        }
        print(f"Returning result: {result}")
        return result
    finally:
        # 删除临时文件
        if Path(temp_file_path).exists():
            Path(temp_file_path).unlink()


def extract_text_from_file(file_path: str) -> str:
    """
    从文件中提取文本内容
    
    Args:
        file_path: 文件路径
        
    Returns:
        文件文本内容
    """
    file_path = Path(file_path)

    if not file_path.exists():
        raise FileNotFoundError(f"文件不存在: {file_path}")

    # 根据文件扩展名选择不同的处理方式
    if file_path.suffix.lower() in ['.txt', '.md']:
        with open(file_path, 'r', encoding='utf-8') as f:
            return f.read()
    elif file_path.suffix.lower() == '.pdf':
        # 需要安装 PyPDF2
        try:
            import PyPDF2
            text = ""
            with open(file_path, 'rb') as f:
                reader = PyPDF2.PdfReader(f)
                for page in reader.pages:
                    text += page.extract_text() + "\n"
            return text
        except ImportError:
            print("警告: 需要安装 PyPDF2 库来解析PDF文件")
            print("运行: pip install PyPDF2")
            return ""
    elif file_path.suffix.lower() in ['.jpg', '.jpeg', '.png']:
        # 使用 cnocr 进行OCR处理
        try:
            # 使用全局OCR实例
            ocr = get_ocr_instance()
            
            # 优化图片以提升OCR效率
            optimized_path = None
            try:
                optimized_path = image_optimization_service.optimize_for_ocr(str(file_path))
                ocr_path = optimized_path
            except Exception as opt_error:
                if settings.DEBUG:
                    print(f"图片优化失败，使用原始图片: {opt_error}")
                ocr_path = str(file_path)
            
            # 进行OCR
            result = ocr.ocr(ocr_path)
            
            # 清理临时优化文件
            if optimized_path and Path(optimized_path).exists():
                try:
                    Path(optimized_path).unlink()
                except:
                    pass
            
            # 提取文本 - 无论det_model设置如何，cnocr都返回字典列表
            if result and isinstance(result, list):
                # 直接连接所有识别出的文本，使用列表推导式提高效率
                text = '\n'.join([item['text'] for item in result if 'text' in item])
            else:
                text = ''
            
            # 只在调试模式下打印详细信息
            if settings.DEBUG:
                print(f"OCR result type: {type(result)}")
                print(f"OCR result length: {len(result) if result else 0}")
                print(f"OCR extracted text: {text[:200] if text else '(empty)'}")
            
            return text
        except ImportError as e:
            print(f"警告: 需要安装 cnocr 库来解析图片文件: {e}")
            print("运行: pip install cnocr")
            return ""
        except Exception as e:
            print(f"警告: OCR处理图片文件失败: {e}")
            import traceback
            traceback.print_exc()
            # 尝试直接返回空字符串，避免因OCR失败导致整个解析过程失败
            return ""
    elif file_path.suffix.lower() in ['.docx']:
        # 需要安装 python-docx
        try:
            from docx import Document
            doc = Document(file_path)
            text = ""
            for para in doc.paragraphs:
                text += para.text + "\n"
            print(f"DOCX extracted text: {text[:100]}...")  # 打印前100个字符
            return text
        except ImportError as e:
            print(f"警告: 需要安装 python-docx 库来解析Word文件: {e}")
            print("运行: pip install python-docx")
            return ""
        except Exception as e:
            print(f"警告: 处理Word文件失败: {e}")
            import traceback
            traceback.print_exc()
            return ""
    elif file_path.suffix.lower() in ['.xlsx', '.xls']:
        # 需要安装 openpyxl
        try:
            import openpyxl
            workbook = openpyxl.load_workbook(file_path)
            text = ""
            for sheet in workbook.sheetnames:
                worksheet = workbook[sheet]
                for row in worksheet.iter_rows(values_only=True):
                    row_text = "\t".join([str(cell) if cell is not None else "" for cell in row])
                    text += row_text + "\n"
            print(f"Excel extracted text: {text[:100]}...")  # 打印前100个字符
            return text
        except ImportError as e:
            print(f"警告: 需要安装 openpyxl 库来解析Excel文件: {e}")
            print("运行: pip install openpyxl")
            return ""
        except Exception as e:
            print(f"警告: 处理Excel文件失败: {e}")
            import traceback
            traceback.print_exc()
            return ""
    elif file_path.suffix.lower() in ['.csv']:
        # 直接读取CSV文件
        try:
            import csv
            text = ""
            # 尝试使用不同的编码读取CSV文件
            encodings = ['utf-8', 'gbk', 'gb2312', 'utf-16']
            for encoding in encodings:
                try:
                    with open(file_path, 'r', encoding=encoding) as f:
                        reader = csv.reader(f)
                        for row in reader:
                            row_text = "\t".join(row)
                            text += row_text + "\n"
                    print(f"CSV extracted text (encoding: {encoding}): {text[:100]}...")  # 打印前100个字符
                    return text
                except UnicodeDecodeError:
                    continue
            # 如果所有编码都失败，使用errors='ignore'模式
            with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
                reader = csv.reader(f)
                for row in reader:
                    row_text = "\t".join(row)
                    text += row_text + "\n"
            print(f"CSV extracted text (encoding: ignore): {text[:100]}...")  # 打印前100个字符
            return text
        except Exception as e:
            print(f"警告: 处理CSV文件失败: {e}")
            import traceback
            traceback.print_exc()
            return ""
    else:
        print(f"警告: 不支持的文件类型: {file_path.suffix}")
        return ""


async def parse_indicators(text: str, report_type: str) -> List[Dict]:
    """
    从文本中解析检查指标
    
    Args:
        text: 检查单文本内容
        report_type: 报告类型 (lab/pathology)
        
    Returns:
        指标列表
    """
    # 使用AI解析检查单文本
    ai_result = await parse_report_with_ai(text, report_type)
    indicators = ai_result.get('indicators', [])
    
    # 确保返回的指标格式正确
    for indicator in indicators:
        # 确保所有必要字段都存在
        if 'name' not in indicator:
            indicator['name'] = ''
        if 'value' not in indicator:
            indicator['value'] = ''
        if 'unit' not in indicator:
            indicator['unit'] = ''
        if 'reference_range' not in indicator:
            indicator['reference_range'] = ''
        if 'is_abnormal' not in indicator:
            indicator['is_abnormal'] = False
    
    return indicators
